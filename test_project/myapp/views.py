from django.http import HttpResponse
from django.shortcuts import render
import json
import openpyxl


def index(request):
    return HttpResponse("MYYY")


def plaintext_read(request):
    with open('plaintext.txt') as f:
        lines = f.readlines()
    return HttpResponse(lines)


def json_read(request):
    with open('test.json') as f:
        data = json.load(f)
    return HttpResponse(data)


def excel_read(request):
    dataframe = openpyxl.load_workbook("data.xlsx")
    dataframe1 = dataframe.active
    my_data = []
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            my_data.append(col[row].value)
    return HttpResponse(my_data)


def html_read(request):
    return render(request, "index.html")
